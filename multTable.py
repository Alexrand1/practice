#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Jan 24 20:39:42 2019

@author: alexrandall
"""

import openpyxl

wb = openpyxl.load_workbook('multiplicationTable.xlsx')
sheet = wb.active


n = int(input('Would you like to see a multiplication table? Import a number '))
i = 1
j = 1
for rowNum in range(2, (n+2)):
    sheet.cell(row=rowNum, column=1).value = i
    i+=1
for colNum in range(2, (n+2)):
    sheet.cell(row=1, column=colNum).value = j
    j+=1
colit = 2
while colit<=n:
    for i in range(2, n+2):
        x = sheet.cell(row=1, column=colit).value
        y = sheet.cell(row=i, column=1).value
        sheet.cell(row=i, column=colit).value = x*y
    colit+=1
    
        

wb.save('multiplicationTable.xlsx')
    