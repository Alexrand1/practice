#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Jan 24 19:43:05 2019

@author: alexrandall
"""

#updatePrices

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {'Garlic' : 3.07,
                 'Celery' : 1.19,
                 'Lemon' : 1.27}

for rowNum in range(2, sheet.max_row): #skip 1st row
    produceName = sheet.cell(row = rowNum, column = 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
w
    